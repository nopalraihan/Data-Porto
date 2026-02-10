"""
ReportGenerator - Generate crosscheck output reports (Excel + PDF).

Produces:
  1. Excel report with color-coded MATCH/MISMATCH/MISSING status
  2. PDF report with verification summary and detailed error log
"""

import os
import unicodedata
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF


def _sanitize(text: str) -> str:
    """Normalize unicode and encode to latin-1 safe string for PDF."""
    text = unicodedata.normalize("NFKD", str(text))
    return text.encode("latin-1", errors="replace").decode("latin-1")


# ──────────────────────────── Color constants ────────────────────────────
MATCH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
MATCH_FONT = Font(color="006100")
MISMATCH_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
MISMATCH_FONT = Font(color="9C0006")
MISSING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
MISSING_FONT = Font(color="9C6500")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


class ReportGenerator:
    """Generate crosscheck result reports."""

    def __init__(self, crosscheck_result: dict, pdf_metadata: dict, output_dir: str):
        self.result = crosscheck_result
        self.pdf_meta = pdf_metadata
        self.output_dir = os.path.abspath(output_dir)
        os.makedirs(self.output_dir, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        doc_name = os.path.splitext(pdf_metadata.get("file_name", "document"))[0]
        self.excel_path = os.path.join(self.output_dir, f"Crosscheck_Result_{doc_name}_{timestamp}.xlsx")
        self.pdf_path = os.path.join(self.output_dir, f"Crosscheck_Result_{doc_name}_{timestamp}.pdf")

    def generate_excel(self) -> str:
        """Generate the crosscheck result Excel report."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Crosscheck Result"

        summary = self.result["summary"]
        results = self.result["results"]

        # ── Title ──
        ws.merge_cells("A1:F1")
        ws.cell(row=1, column=1, value="PLN DOCUMENT CROSSCHECK REPORT").font = Font(
            bold=True, size=14, color="1F4E79"
        )
        ws.cell(row=2, column=1, value=f"Document: {self.pdf_meta.get('file_name', 'N/A')}").font = Font(size=10)
        ws.cell(row=2, column=4, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = Font(
            size=10, color="666666"
        )

        # ── Verification Status Banner ──
        row = 4
        ws.merge_cells(f"A{row}:F{row}")
        pct = summary["match_percentage"]
        if pct == 100:
            status_text = "VERIFIED - All fields match"
            status_fill = MATCH_FILL
            status_font = Font(bold=True, size=14, color="006100")
        elif pct >= 70:
            status_text = f"PARTIAL MATCH - {pct}% fields match ({summary['total_mismatch']} mismatches)"
            status_fill = MISSING_FILL
            status_font = Font(bold=True, size=14, color="9C6500")
        else:
            status_text = f"MISMATCH - Only {pct}% fields match ({summary['total_mismatch']} errors found)"
            status_fill = MISMATCH_FILL
            status_font = Font(bold=True, size=14, color="9C0006")

        cell = ws.cell(row=row, column=1, value=status_text)
        cell.font = status_font
        cell.fill = status_fill
        cell.alignment = Alignment(horizontal="center")
        for c in range(1, 7):
            ws.cell(row=row, column=c).fill = status_fill

        # ── Summary Stats ──
        row = 6
        stats = [
            ("Matched Row #", summary.get("matched_row_number", "N/A")),
            ("Fields Checked", summary["total_fields_checked"]),
            ("Matches", summary["total_match"]),
            ("Mismatches", summary["total_mismatch"]),
            ("Missing", summary["total_missing"]),
            ("Match %", f"{pct}%"),
        ]
        for label, value in stats:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
            ws.cell(row=row, column=2, value=str(value)).font = Font(size=10)
            row += 1

        # ── Detail Table ──
        row += 1
        headers = ["Field Name", "PDF Value", "Excel Value", "Status", "Notes"]
        widths = [22, 30, 30, 14, 40]

        for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        header_row = row
        row += 1

        for r in results:
            status = r["match_status"]
            fill = MATCH_FILL if status == "MATCH" else (MISMATCH_FILL if status == "MISMATCH" else MISSING_FILL)
            font = MATCH_FONT if status == "MATCH" else (MISMATCH_FONT if status == "MISMATCH" else MISSING_FONT)

            values = [r["field_name"], r.get("pdf_value", ""), r.get("excel_value", ""), status, r.get("notes", "")]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_idx, value=val or "")
                cell.border = BORDER
                cell.font = font if col_idx == 4 else Font(size=10)
                if col_idx == 4:
                    cell.fill = fill
                    cell.alignment = Alignment(horizontal="center")
            row += 1

        # Auto-filter on detail table
        ws.auto_filter.ref = f"A{header_row}:E{row - 1}"

        # ── Error Log Sheet (only mismatches) ──
        mismatches = [r for r in results if r["match_status"] == "MISMATCH"]
        if mismatches:
            ws2 = wb.create_sheet("Error Log")
            ws2.cell(row=1, column=1, value="MISMATCH ERROR LOG").font = Font(bold=True, size=14, color="9C0006")
            ws2.cell(row=2, column=1, value=f"Document: {self.pdf_meta.get('file_name', '')}").font = Font(size=10)
            ws2.cell(row=2, column=3, value=f"Total Errors: {len(mismatches)}").font = Font(
                bold=True, size=10, color="9C0006"
            )

            err_headers = ["#", "Field Name", "PDF Value", "Excel Value", "Notes"]
            err_widths = [6, 22, 30, 30, 45]
            for col_idx, (h, w) in enumerate(zip(err_headers, err_widths), start=1):
                cell = ws2.cell(row=4, column=col_idx, value=h)
                cell.font = HEADER_FONT
                cell.fill = PatternFill(start_color="9C0006", end_color="9C0006", fill_type="solid")
                cell.border = BORDER
                ws2.column_dimensions[get_column_letter(col_idx)].width = w

            for i, r in enumerate(mismatches, start=1):
                rr = i + 4
                ws2.cell(row=rr, column=1, value=i).border = BORDER
                ws2.cell(row=rr, column=2, value=r["field_name"]).border = BORDER
                ws2.cell(row=rr, column=3, value=r.get("pdf_value", "")).border = BORDER
                ws2.cell(row=rr, column=3).fill = MISMATCH_FILL
                ws2.cell(row=rr, column=4, value=r.get("excel_value", "")).border = BORDER
                ws2.cell(row=rr, column=4).fill = MISMATCH_FILL
                ws2.cell(row=rr, column=5, value=r.get("notes", "")).border = BORDER

        # ── ML Analysis Sheet ──
        if "ml_confidence" in self.result or "ml_anomalies" in self.result:
            ws_ml = wb.create_sheet("ML Analysis")
            ML_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            ws_ml.cell(row=1, column=1, value="ML-POWERED ANALYSIS").font = Font(bold=True, size=14, color="1F4E79")

            r = 3
            # Confidence score
            if "ml_confidence" in self.result:
                conf = self.result["ml_confidence"]
                ws_ml.cell(row=r, column=1, value="Random Forest Confidence Score").font = Font(bold=True, size=12, color="1F4E79")
                r += 1
                for label, val in [
                    ("Confidence Score", f"{conf['confidence_score']}%"),
                    ("Prediction", conf["prediction"]),
                    ("Risk Level", conf["risk_level"]),
                ]:
                    ws_ml.cell(row=r, column=1, value=label).font = Font(bold=True, size=10)
                    cell = ws_ml.cell(row=r, column=2, value=val)
                    cell.font = Font(bold=True, size=10)
                    if val == "VALID":
                        cell.font = Font(bold=True, color="006100")
                    elif val in ("SUSPICIOUS", "HIGH"):
                        cell.font = Font(bold=True, color="9C0006")
                    r += 1

                r += 1
                ws_ml.cell(row=r, column=1, value="Feature Contributions").font = Font(bold=True, size=11, color="1F4E79")
                r += 1
                for h_idx, h in enumerate(["Feature", "Weight"], start=1):
                    cell = ws_ml.cell(row=r, column=h_idx, value=h)
                    cell.font = HEADER_FONT
                    cell.fill = ML_HEADER
                    cell.border = BORDER
                r += 1
                for feat, weight in conf["feature_contributions"].items():
                    ws_ml.cell(row=r, column=1, value=feat).border = BORDER
                    ws_ml.cell(row=r, column=2, value=round(weight, 4)).border = BORDER
                    ws_ml.cell(row=r, column=2).number_format = "0.0000"
                    r += 1

            # Text similarity
            if "ml_similarity" in self.result:
                r += 1
                ws_ml.cell(row=r, column=1, value="TF-IDF Text Similarity").font = Font(bold=True, size=12, color="1F4E79")
                r += 1
                for h_idx, h in enumerate(["Field", "Score", "Classification"], start=1):
                    cell = ws_ml.cell(row=r, column=h_idx, value=h)
                    cell.font = HEADER_FONT
                    cell.fill = ML_HEADER
                    cell.border = BORDER
                r += 1
                for field, sim in self.result["ml_similarity"].items():
                    ws_ml.cell(row=r, column=1, value=field).border = BORDER
                    ws_ml.cell(row=r, column=2, value=sim["score"]).border = BORDER
                    ws_ml.cell(row=r, column=2).number_format = "0.0000"
                    ws_ml.cell(row=r, column=3, value=sim["classification"]).border = BORDER
                    r += 1

            # Anomaly flags
            if "ml_anomalies" in self.result:
                flags = self.result["ml_anomalies"]
                r += 1
                ws_ml.cell(row=r, column=1, value=f"Anomaly Detection ({len(flags)} flags)").font = Font(
                    bold=True, size=12, color="1F4E79"
                )
                r += 1
                if flags:
                    for h_idx, h in enumerate(["Severity", "Field", "Message", "Expected", "Actual"], start=1):
                        cell = ws_ml.cell(row=r, column=h_idx, value=h)
                        cell.font = HEADER_FONT
                        cell.fill = ML_HEADER
                        cell.border = BORDER
                    r += 1
                    for f in flags:
                        ws_ml.cell(row=r, column=1, value=f["severity"]).border = BORDER
                        ws_ml.cell(row=r, column=1).fill = MISMATCH_FILL if f["severity"] == "CRITICAL" else MISSING_FILL
                        ws_ml.cell(row=r, column=2, value=f["field"]).border = BORDER
                        ws_ml.cell(row=r, column=3, value=f["message"]).border = BORDER
                        ws_ml.cell(row=r, column=4, value=f["expected"]).border = BORDER
                        ws_ml.cell(row=r, column=5, value=f["actual"]).border = BORDER
                        r += 1
                else:
                    ws_ml.cell(row=r, column=1, value="No anomalies detected.").font = Font(italic=True, color="006100")

            for c in range(1, 6):
                ws_ml.column_dimensions[get_column_letter(c)].width = [30, 18, 18, 25, 25][c-1]

        wb.save(self.excel_path)
        return self.excel_path

    def generate_pdf(self) -> str:
        """Generate the crosscheck result PDF report."""
        summary = self.result["summary"]
        results = self.result["results"]
        pct = summary["match_percentage"]

        pdf = FPDF(orientation="P", unit="mm", format="A4")
        pdf.set_auto_page_break(auto=True, margin=20)
        page_w = pdf.w - 2 * pdf.l_margin

        # ── Title Page ──
        pdf.add_page()
        pdf.set_y(60)
        pdf.set_font("Helvetica", "B", 24)
        pdf.set_text_color(31, 78, 121)
        pdf.cell(0, 12, _sanitize("PLN Document Crosscheck Report"), align="C", new_x="LMARGIN", new_y="NEXT")

        pdf.set_font("Helvetica", "", 12)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(0, 8, _sanitize(f"Document: {self.pdf_meta.get('file_name', 'N/A')}"), align="C",
                 new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 8, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", align="C",
                 new_x="LMARGIN", new_y="NEXT")

        pdf.ln(15)

        # Verdict banner
        if pct == 100:
            pdf.set_fill_color(198, 239, 206)
            pdf.set_text_color(0, 97, 0)
            verdict = "VERIFIED - Document is VALID"
        elif pct >= 70:
            pdf.set_fill_color(255, 235, 156)
            pdf.set_text_color(156, 101, 0)
            verdict = f"PARTIAL MATCH - {pct}% match"
        else:
            pdf.set_fill_color(255, 199, 206)
            pdf.set_text_color(156, 0, 6)
            verdict = f"MISMATCH - Only {pct}% match"

        pdf.set_font("Helvetica", "B", 18)
        pdf.cell(0, 14, _sanitize(verdict), align="C", fill=True, new_x="LMARGIN", new_y="NEXT")

        # ── Summary Page ──
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 16)
        pdf.set_text_color(31, 78, 121)
        pdf.cell(0, 10, "Verification Summary", new_x="LMARGIN", new_y="NEXT")
        pdf.set_draw_color(31, 78, 121)
        pdf.line(pdf.l_margin, pdf.get_y(), pdf.l_margin + page_w, pdf.get_y())
        pdf.ln(5)

        kv_items = [
            ("Document", self.pdf_meta.get("file_name", "N/A")),
            ("Pages", str(self.pdf_meta.get("page_count", "N/A"))),
            ("Matched Excel Row", str(summary.get("matched_row_number", "N/A"))),
            ("Fields Checked", str(summary["total_fields_checked"])),
            ("Matches", str(summary["total_match"])),
            ("Mismatches", str(summary["total_mismatch"])),
            ("Missing", str(summary["total_missing"])),
            ("Match Percentage", f"{pct}%"),
        ]

        pdf.set_text_color(0, 0, 0)
        for key, val in kv_items:
            pdf.set_font("Helvetica", "B", 10)
            pdf.cell(55, 7, _sanitize(f"{key}:"), new_x="END")
            pdf.set_font("Helvetica", "", 10)
            pdf.cell(0, 7, _sanitize(val), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(5)

        # ── Detail Table ──
        pdf.set_font("Helvetica", "B", 14)
        pdf.set_text_color(31, 78, 121)
        pdf.cell(0, 10, "Field-by-Field Comparison", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

        col_widths = [40, 42, 42, 22, 44]
        headers = ["Field", "PDF Value", "Excel Value", "Status", "Notes"]

        pdf.set_font("Helvetica", "B", 8)
        pdf.set_fill_color(31, 78, 121)
        pdf.set_text_color(255, 255, 255)
        for i, h in enumerate(headers):
            pdf.cell(col_widths[i], 7, h, border=1, fill=True, align="C", new_x="END")
        pdf.ln()

        pdf.set_font("Helvetica", "", 7)
        for row_idx, r in enumerate(results):
            status = r["match_status"]
            if status == "MATCH":
                pdf.set_fill_color(198, 239, 206)
            elif status == "MISMATCH":
                pdf.set_fill_color(255, 199, 206)
            else:
                pdf.set_fill_color(255, 235, 156)

            pdf.set_text_color(0, 0, 0)
            values = [
                r["field_name"],
                (r.get("pdf_value") or "—")[:30],
                (r.get("excel_value") or "—")[:30],
                status,
                (r.get("notes") or "")[:35],
            ]
            for i, val in enumerate(values):
                fill = True if i == 3 else (row_idx % 2 == 1)
                if i != 3 and row_idx % 2 == 1:
                    pdf.set_fill_color(242, 242, 242)
                elif i != 3:
                    pdf.set_fill_color(255, 255, 255)

                if i == 3:
                    if status == "MATCH":
                        pdf.set_fill_color(198, 239, 206)
                    elif status == "MISMATCH":
                        pdf.set_fill_color(255, 199, 206)
                    else:
                        pdf.set_fill_color(255, 235, 156)

                pdf.cell(col_widths[i], 6, _sanitize(val), border=1, fill=True, align="C" if i == 3 else "L",
                         new_x="END")
            pdf.ln()

        # ── Error Log Section ──
        mismatches = [r for r in results if r["match_status"] == "MISMATCH"]
        if mismatches:
            pdf.add_page()
            pdf.set_font("Helvetica", "B", 16)
            pdf.set_text_color(156, 0, 6)
            pdf.cell(0, 10, f"Error Log ({len(mismatches)} Mismatches)", new_x="LMARGIN", new_y="NEXT")
            pdf.set_draw_color(156, 0, 6)
            pdf.line(pdf.l_margin, pdf.get_y(), pdf.l_margin + page_w, pdf.get_y())
            pdf.ln(5)

            for i, r in enumerate(mismatches, start=1):
                pdf.set_font("Helvetica", "B", 10)
                pdf.set_text_color(156, 0, 6)
                pdf.cell(0, 7, _sanitize(f"Error #{i}: {r['field_name']}"), new_x="LMARGIN", new_y="NEXT")

                pdf.set_font("Helvetica", "", 9)
                pdf.set_text_color(0, 0, 0)
                pdf.cell(35, 6, "PDF Value:", new_x="END")
                pdf.set_font("Helvetica", "B", 9)
                pdf.cell(0, 6, _sanitize(r.get("pdf_value") or "N/A"), new_x="LMARGIN", new_y="NEXT")

                pdf.set_font("Helvetica", "", 9)
                pdf.cell(35, 6, "Excel Value:", new_x="END")
                pdf.set_font("Helvetica", "B", 9)
                pdf.cell(0, 6, _sanitize(r.get("excel_value") or "N/A"), new_x="LMARGIN", new_y="NEXT")

                pdf.set_font("Helvetica", "", 9)
                pdf.cell(35, 6, "Notes:", new_x="END")
                pdf.cell(0, 6, _sanitize(r.get("notes") or ""), new_x="LMARGIN", new_y="NEXT")
                pdf.ln(4)

        # ── ML Analysis Section ──
        has_ml = any(k in self.result for k in ("ml_confidence", "ml_similarity", "ml_anomalies"))
        if has_ml:
            pdf.add_page()
            pdf.set_font("Helvetica", "B", 16)
            pdf.set_text_color(31, 78, 121)
            pdf.cell(0, 10, "ML-Powered Analysis", new_x="LMARGIN", new_y="NEXT")
            pdf.set_draw_color(31, 78, 121)
            pdf.line(pdf.l_margin, pdf.get_y(), pdf.l_margin + page_w, pdf.get_y())
            pdf.ln(5)

            # Confidence score
            if "ml_confidence" in self.result:
                conf = self.result["ml_confidence"]
                pdf.set_font("Helvetica", "B", 13)
                pdf.set_text_color(31, 78, 121)
                pdf.cell(0, 8, "Random Forest Confidence Scorer", new_x="LMARGIN", new_y="NEXT")
                pdf.ln(2)

                score_val = conf["confidence_score"]
                if score_val >= 85:
                    pdf.set_fill_color(198, 239, 206)
                elif score_val >= 60:
                    pdf.set_fill_color(255, 235, 156)
                else:
                    pdf.set_fill_color(255, 199, 206)

                pdf.set_font("Helvetica", "B", 20)
                pdf.set_text_color(0, 0, 0)
                pdf.cell(0, 12, _sanitize(f"{score_val}% Confidence  |  {conf['prediction']}  |  Risk: {conf['risk_level']}"),
                         fill=True, align="C", new_x="LMARGIN", new_y="NEXT")
                pdf.ln(5)

                # Feature weights table
                pdf.set_font("Helvetica", "B", 8)
                pdf.set_fill_color(68, 114, 196)
                pdf.set_text_color(255, 255, 255)
                pdf.cell(60, 6, "Feature", border=1, fill=True, align="C", new_x="END")
                pdf.cell(30, 6, "Weight", border=1, fill=True, align="C", new_x="LMARGIN", new_y="NEXT")
                pdf.set_font("Helvetica", "", 8)
                pdf.set_text_color(0, 0, 0)
                for feat, weight in conf["feature_contributions"].items():
                    pdf.set_fill_color(242, 242, 242)
                    pdf.cell(60, 5, _sanitize(feat), border=1, fill=True, new_x="END")
                    pdf.cell(30, 5, f"{weight:+.4f}", border=1, fill=True, align="C", new_x="LMARGIN", new_y="NEXT")
                pdf.ln(5)

            # Text similarity
            if "ml_similarity" in self.result:
                pdf.set_font("Helvetica", "B", 13)
                pdf.set_text_color(31, 78, 121)
                pdf.cell(0, 8, "TF-IDF Text Similarity", new_x="LMARGIN", new_y="NEXT")
                pdf.ln(2)

                pdf.set_font("Helvetica", "B", 8)
                pdf.set_fill_color(68, 114, 196)
                pdf.set_text_color(255, 255, 255)
                for h, w in [("Field", 50), ("Score", 25), ("Classification", 35)]:
                    pdf.cell(w, 6, h, border=1, fill=True, align="C", new_x="END")
                pdf.ln()

                pdf.set_font("Helvetica", "", 8)
                pdf.set_text_color(0, 0, 0)
                for field, sim in self.result["ml_similarity"].items():
                    pdf.cell(50, 5, _sanitize(field), border=1, new_x="END")
                    pdf.cell(25, 5, f"{sim['score']:.4f}", border=1, align="C", new_x="END")
                    pdf.cell(35, 5, sim["classification"], border=1, align="C", new_x="LMARGIN", new_y="NEXT")
                pdf.ln(5)

            # Anomalies
            if "ml_anomalies" in self.result:
                flags = self.result["ml_anomalies"]
                pdf.set_font("Helvetica", "B", 13)
                pdf.set_text_color(31, 78, 121)
                pdf.cell(0, 8, _sanitize(f"Anomaly Detection ({len(flags)} flags)"), new_x="LMARGIN", new_y="NEXT")
                pdf.ln(2)

                if flags:
                    for f in flags:
                        sev_color = (156, 0, 6) if f["severity"] == "CRITICAL" else (156, 101, 0)
                        pdf.set_font("Helvetica", "B", 9)
                        pdf.set_text_color(*sev_color)
                        pdf.cell(0, 6, _sanitize(f"[{f['severity']}] {f['field']}: {f['message']}"),
                                 new_x="LMARGIN", new_y="NEXT")
                        pdf.set_font("Helvetica", "", 8)
                        pdf.set_text_color(0, 0, 0)
                        pdf.cell(0, 5, _sanitize(f"  Expected: {f['expected']}  |  Actual: {f['actual']}"),
                                 new_x="LMARGIN", new_y="NEXT")
                        pdf.ln(2)
                else:
                    pdf.set_font("Helvetica", "I", 10)
                    pdf.set_text_color(0, 97, 0)
                    pdf.cell(0, 6, "No anomalies detected.", new_x="LMARGIN", new_y="NEXT")

        # Footer
        total_pages = pdf.pages_count
        for i in range(1, total_pages + 1):
            pdf.page = i
            pdf.set_y(-15)
            pdf.set_font("Helvetica", "I", 8)
            pdf.set_text_color(150, 150, 150)
            pdf.cell(0, 10, f"Page {i} of {total_pages} | PLN Crosscheck Report", align="C")

        pdf.output(self.pdf_path)
        return self.pdf_path

    def generate_all(self) -> dict:
        """Generate both Excel and PDF reports. Return paths."""
        return {
            "excel": self.generate_excel(),
            "pdf": self.generate_pdf(),
        }
