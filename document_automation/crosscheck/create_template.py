"""
Creates the PLN Crosscheck Excel template with the expected structure.
Run this script to generate a fresh template file.
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def create_template(output_path: str, sample_data: bool = True) -> str:
    """
    Create the PLN_Crosscheck_Template.xlsx file.

    Columns match PLN document fields for 1:1 crosschecking:
      - id_pelanggan, nama_pelanggan, alamat, tarif_daya,
        nomor_meter, nomor_kwh, periode,
        stand_meter_awal, stand_meter_akhir, pemakaian_kwh,
        biaya_listrik, status
    """
    wb = Workbook()

    # --- Sheet 1: Data Pelanggan (main crosscheck sheet) ---
    ws = wb.active
    ws.title = "Data Pelanggan"

    HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    columns = [
        ("No", 6),
        ("ID Pelanggan", 18),
        ("Nama Pelanggan", 25),
        ("Alamat", 35),
        ("Tarif/Daya", 15),
        ("Nomor Meter", 18),
        ("Nomor kWh", 18),
        ("Periode", 15),
        ("Stand Meter Awal", 18),
        ("Stand Meter Akhir", 18),
        ("Pemakaian (kWh)", 18),
        ("Biaya Listrik (Rp)", 22),
        ("Status", 15),
    ]

    # Title row
    ws.merge_cells("A1:M1")
    title_cell = ws.cell(row=1, column=1, value="PLN - DATA CROSSCHECK PELANGGAN")
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:M2")
    ws.cell(row=2, column=1, value="Template untuk verifikasi data dokumen PLN").font = Font(
        name="Calibri", italic=True, size=10, color="666666"
    )
    ws["A2"].alignment = Alignment(horizontal="center")

    # Headers at row 4
    header_row = 4
    for col_idx, (col_name, width) in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Status dropdown validation
    status_dv = DataValidation(
        type="list",
        formula1='"Verified,Pending,Mismatch"',
        allow_blank=True,
    )
    status_dv.error = "Pilih status: Verified, Pending, atau Mismatch"
    status_dv.errorTitle = "Status tidak valid"
    ws.add_data_validation(status_dv)

    # Sample data rows (realistic PLN data based on Penggilingan Elok area)
    sample_rows = []
    if sample_data:
        sample_rows = [
            [1, "532100012345", "SUHARTO", "JL. PENGGILINGAN ELOK NO.23 RT005/RW012, PENGGILINGAN, CAKUNG, JAKARTA TIMUR",
             "R1/1300 VA", "JTX476", "85201234", "Januari 2026", "15230", "15480", "250", "352500", "Pending"],
            [2, "532100012346", "DEWI SARTIKA", "JL. PENGGILINGAN ELOK NO.25 RT005/RW012, PENGGILINGAN, CAKUNG, JAKARTA TIMUR",
             "R1/2200 VA", "JTX477", "85201235", "Januari 2026", "22100", "22450", "350", "493500", "Pending"],
            [3, "532100012347", "BAMBANG WIJAYA", "JL. PENGGILINGAN ELOK NO.27 RT005/RW012, PENGGILINGAN, CAKUNG, JAKARTA TIMUR",
             "R1M/900 VA", "JTX478", "85201236", "Januari 2026", "8050", "8200", "150", "211500", "Pending"],
            [4, "532100012348", "RATNA SARI", "JL. PENGGILINGAN ELOK NO.29 RT005/RW012, PENGGILINGAN, CAKUNG, JAKARTA TIMUR",
             "R1/1300 VA", "JTX479", "85201237", "Januari 2026", "12500", "12780", "280", "394800", "Pending"],
            [5, "532100012349", "AHMAD FAUZI", "JL. PENGGILINGAN ELOK NO.31 RT005/RW012, PENGGILINGAN, CAKUNG, JAKARTA TIMUR",
             "R2/3500 VA", "JTX480", "85201238", "Januari 2026", "45600", "46100", "500", "705000", "Pending"],
        ]

    for row_idx, row_data in enumerate(sample_rows, start=header_row + 1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Calibri", size=10)
            cell.border = BORDER
            if col_idx in (9, 10, 11):  # numeric meter fields
                cell.number_format = "#,##0"
            elif col_idx == 12:  # biaya
                cell.number_format = "#,##0"
        # Apply status validation
        status_dv.add(ws.cell(row=row_idx, column=13))

    # Add validation for 50 more potential rows
    for r in range(header_row + len(sample_rows) + 1, header_row + 55):
        status_dv.add(ws.cell(row=r, column=13))

    # Freeze panes
    ws.freeze_panes = f"A{header_row + 1}"
    # Auto-filter
    ws.auto_filter.ref = f"A{header_row}:M{header_row + max(len(sample_rows), 1)}"

    # --- Sheet 2: Field Mapping Reference ---
    ws2 = wb.create_sheet("Field Mapping")
    ws2.cell(row=1, column=1, value="Field Mapping Reference").font = Font(bold=True, size=12, color="1F4E79")

    mapping_headers = ["Excel Column", "PDF Field Key", "Description", "Example"]
    mapping_data = [
        ["ID Pelanggan", "id_pelanggan", "12-digit PLN customer ID", "532100012345"],
        ["Nama Pelanggan", "nama_pelanggan", "Customer name as in PLN records", "SUHARTO"],
        ["Alamat", "alamat", "Full address from PLN document", "JL. PENGGILINGAN ELOK NO.23..."],
        ["Tarif/Daya", "tarif_daya", "Tariff group / connected power", "R1/1300 VA"],
        ["Nomor Meter", "nomor_meter", "Meter device number", "JTX476"],
        ["Nomor kWh", "nomor_kwh", "kWh meter serial number", "85201234"],
        ["Periode", "periode", "Billing period", "Januari 2026"],
        ["Stand Meter Awal", "stand_meter_awal", "Starting meter reading", "15230"],
        ["Stand Meter Akhir", "stand_meter_akhir", "Ending meter reading", "15480"],
        ["Pemakaian (kWh)", "pemakaian_kwh", "Total consumption in kWh", "250"],
        ["Biaya Listrik (Rp)", "biaya_listrik", "Total bill amount in Rupiah", "352500"],
        ["Status", "status", "Verification status", "Verified / Pending / Mismatch"],
    ]

    for col_idx, header in enumerate(mapping_headers, start=1):
        cell = ws2.cell(row=3, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER

    for row_idx, row_data in enumerate(mapping_data, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER

    for col_idx, width in enumerate([20, 20, 40, 30], start=1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return os.path.abspath(output_path)


if __name__ == "__main__":
    path = create_template(
        os.path.join(os.path.dirname(__file__), "..", "sample_data", "PLN_Crosscheck_Template.xlsx")
    )
    print(f"Template created: {path}")
