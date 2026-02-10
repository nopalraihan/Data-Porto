"""
ExcelWriter - Write DataFrames to formatted Excel workbooks.
"""

import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


class ExcelWriter:
    """Write DataFrames to Excel files with professional formatting."""

    # Style constants
    HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    CELL_FONT = Font(name="Calibri", size=10)
    CELL_ALIGNMENT = Alignment(vertical="center")
    BORDER = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2F5496")
    SUBTITLE_FONT = Font(name="Calibri", size=10, color="666666")

    def __init__(self, output_path: str):
        self.output_path = os.path.abspath(output_path)
        self.wb = Workbook()
        # Remove the default sheet
        self.wb.remove(self.wb.active)

    def add_dataframe_sheet(
        self,
        df: pd.DataFrame,
        sheet_name: str = "Data",
        title: str = None,
        include_chart: bool = False,
        chart_col: str = None,
        chart_label_col: str = None,
    ) -> None:
        """Add a DataFrame as a formatted sheet in the workbook."""
        ws = self.wb.create_sheet(title=sheet_name)

        current_row = 1

        # Title and subtitle
        if title:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
            title_cell = ws.cell(row=1, column=1, value=title)
            title_cell.font = self.TITLE_FONT
            title_cell.alignment = Alignment(horizontal="left")

            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(df.columns))
            sub_cell = ws.cell(row=2, column=1, value=f"Generated: {timestamp}")
            sub_cell.font = self.SUBTITLE_FONT
            current_row = 4

        # Write headers
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=str(col_name))
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.BORDER

        header_row = current_row
        current_row += 1

        # Write data rows
        for row_idx, (_, row) in enumerate(df.iterrows()):
            for col_idx, value in enumerate(row, start=1):
                # Convert numpy types to native Python types
                if hasattr(value, "item"):
                    value = value.item()
                elif pd.isna(value):
                    value = None

                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.font = self.CELL_FONT
                cell.alignment = self.CELL_ALIGNMENT
                cell.border = self.BORDER

                # Alternating row colors
                if row_idx % 2 == 1:
                    cell.fill = self.ALT_ROW_FILL

                # Number formatting
                if isinstance(value, float):
                    cell.number_format = "#,##0.00"
                elif isinstance(value, int):
                    cell.number_format = "#,##0"

            current_row += 1

        # Auto-fit column widths
        for col_idx in range(1, len(df.columns) + 1):
            max_length = len(str(df.columns[col_idx - 1]))
            for row in ws.iter_rows(min_row=header_row, max_row=current_row - 1, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 4, 50)

        # Auto-filter
        last_col = get_column_letter(len(df.columns))
        ws.auto_filter.ref = f"A{header_row}:{last_col}{current_row - 1}"

        # Freeze header row
        ws.freeze_panes = f"A{header_row + 1}"

        # Optional bar chart
        if include_chart and chart_col and chart_label_col:
            self._add_bar_chart(ws, df, header_row, current_row - 1, chart_col, chart_label_col)

    def add_summary_sheet(
        self,
        data_df: pd.DataFrame,
        stats_df: pd.DataFrame,
        column_info_df: pd.DataFrame,
        sheet_name: str = "Summary",
    ) -> None:
        """Add a summary/overview sheet with statistics and column info."""
        ws = self.wb.create_sheet(title=sheet_name)

        # Title
        ws.merge_cells("A1:F1")
        title_cell = ws.cell(row=1, column=1, value="Data Summary Report")
        title_cell.font = self.TITLE_FONT

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.merge_cells("A2:F2")
        ws.cell(row=2, column=1, value=f"Generated: {timestamp}").font = self.SUBTITLE_FONT

        # Overview section
        row = 4
        ws.cell(row=row, column=1, value="Overview").font = Font(bold=True, size=12, color="2F5496")
        row += 1
        overview_items = [
            ("Total Rows", len(data_df)),
            ("Total Columns", len(data_df.columns)),
            ("Numeric Columns", len(data_df.select_dtypes(include="number").columns)),
            ("Text Columns", len(data_df.select_dtypes(include="object").columns)),
            ("Total Null Values", int(data_df.isnull().sum().sum())),
            ("Memory Usage (KB)", round(data_df.memory_usage(deep=True).sum() / 1024, 2)),
        ]
        for label, value in overview_items:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
            ws.cell(row=row, column=2, value=value).font = self.CELL_FONT
            row += 1

        # Column Info
        row += 1
        ws.cell(row=row, column=1, value="Column Information").font = Font(bold=True, size=12, color="2F5496")
        row += 1
        self._write_mini_table(ws, column_info_df, row)
        row += len(column_info_df) + 2

        # Statistics
        if not stats_df.empty:
            row += 1
            ws.cell(row=row, column=1, value="Descriptive Statistics").font = Font(bold=True, size=12, color="2F5496")
            row += 1
            stats_reset = stats_df.reset_index()
            stats_reset.rename(columns={"index": "Statistic"}, inplace=True)
            self._write_mini_table(ws, stats_reset, row)

        # Auto-fit columns
        for col_idx in range(1, 10):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

    def _write_mini_table(self, ws, df: pd.DataFrame, start_row: int) -> None:
        """Write a small formatted table into a worksheet."""
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=str(col_name))
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.border = self.BORDER

        for row_idx, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
            for col_idx, value in enumerate(row, start=1):
                if hasattr(value, "item"):
                    value = value.item()
                elif pd.isna(value):
                    value = None
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.CELL_FONT
                cell.border = self.BORDER
                if isinstance(value, float):
                    cell.number_format = "#,##0.00"

    def _add_bar_chart(self, ws, df, header_row, last_data_row, value_col, label_col):
        """Insert a bar chart into the sheet."""
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = f"{value_col} by {label_col}"
        chart.y_axis.title = value_col
        chart.x_axis.title = label_col

        value_col_idx = list(df.columns).index(value_col) + 1
        label_col_idx = list(df.columns).index(label_col) + 1

        data_ref = Reference(ws, min_col=value_col_idx, min_row=header_row, max_row=last_data_row)
        cats_ref = Reference(ws, min_col=label_col_idx, min_row=header_row + 1, max_row=last_data_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        chart.width = 20
        chart.height = 12

        chart_col = get_column_letter(len(df.columns) + 2)
        ws.add_chart(chart, f"{chart_col}{header_row}")

    def save(self) -> str:
        """Save the workbook and return the output path."""
        os.makedirs(os.path.dirname(self.output_path), exist_ok=True)
        self.wb.save(self.output_path)
        return self.output_path
